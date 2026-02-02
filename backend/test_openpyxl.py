import openpyxl
import os

path = "test_write.xlsx"
try:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = "Test"
    wb.save(path)
    print("Successfully saved test file")
    
    wb2 = openpyxl.load_workbook(path)
    print("Successfully loaded test file")
    print("Cell A1:", wb2.active['A1'].value)
except Exception as e:
    print(f"Environment test failed: {e}")
finally:
    if os.path.exists(path):
        os.remove(path)
