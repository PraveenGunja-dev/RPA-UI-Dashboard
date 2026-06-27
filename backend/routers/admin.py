from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks
from sqlalchemy.orm import Session
from fastapi.responses import FileResponse, StreamingResponse
from typing import List, Optional
from pydantic import BaseModel
import os
import json
import re
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
from mail_util import send_new_bot_notification

from database import get_db
from models import FileLog, Bot, Department, SPOC, RegisteredUser, AuditLog
from .auth import get_current_user_email


router = APIRouter(
    prefix="/api/admin",
    tags=["admin"]
)

# Pydantic schemas for bot management
class BotCreate(BaseModel):
    use_case_name: str
    use_case_no: Optional[str] = None
    bot_name: Optional[str] = None
    department_id: Optional[int] = None
    spoc_id: Optional[int] = None
    status: Optional[str] = "Active"
    developer: Optional[str] = None
    hours_saved_monthly: Optional[float] = 0
    pdd_link: Optional[str] = None
    schedule_time: Optional[str] = None
    description: Optional[str] = None

class BotUpdate(BotCreate):
    pass

class SPOCOut(BaseModel):
    id: int
    name: str
    
    class Config:
        from_attributes = True

class UserCreate(BaseModel):
    email: str
    name: Optional[str] = None
    role: Optional[str] = "User"

class UserUpdate(UserCreate):
    is_active: Optional[int] = 1

class UserOut(BaseModel):
    id: int
    email: str
    name: Optional[str]
    role: str
    is_active: int
    
    class Config:
        from_attributes = True

# File Logs endpoints
@router.get("/logs")
def get_file_logs(db: Session = Depends(get_db)):
    """Fetch all file process logs."""
    logs = db.query(FileLog).order_by(FileLog.file_date.desc(), FileLog.id.desc()).all()
    return logs

@router.get("/download/{log_id}")
def download_file(log_id: int, db: Session = Depends(get_db)):
    """Download the archived Excel file."""
    log = db.query(FileLog).filter(FileLog.id == log_id).first()
    if not log or not log.file_path:
        raise HTTPException(status_code=404, detail="File log not found or file not available")
    
    abs_path = os.path.join(os.getcwd(), log.file_path)
    
    if not os.path.exists(abs_path):
        raise HTTPException(status_code=404, detail="Physical file not found")
        
    return FileResponse(
        path=abs_path,
        filename=os.path.basename(abs_path),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# SPOC endpoints
@router.get("/spocs", response_model=List[SPOCOut])
def get_all_spocs(db: Session = Depends(get_db)):
    """Get all SPOCs for dropdown selection."""
    spocs = db.query(SPOC).order_by(SPOC.name).all()
    return spocs

@router.get("/export-bots")
def export_bots(db: Session = Depends(get_db)):
    """Export all bots to a beautifully formatted Excel report."""
    bots = db.query(Bot).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Bot Report"
    
    # Hide gridlines for a clean look
    ws.sheet_view.showGridLines = False
    
    # Headers and Title
    headers = [
        "S. No.", "Category", "CoBot Name", "Status", "Description Available",
        "PDD Available", "Description", "Hours Saved (Monthly)", "Hours Saved (This Month)", 
        "Runs This Month", "Remarks"
    ]
    
    # Add Heading
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws['A1']
    title_cell.value = "CoBot Report"
    title_cell.font = Font(size=18, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    
    # Fixed Column Widths
    col_widths = {
        'A': 8, 'B': 20, 'C': 35, 'D': 15, 'E': 20,
        'F': 15, 'G': 50, 'H': 22, 'I': 22, 'J': 15, 'K': 30
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
        
    # Write Headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                    top=Side(style='thin'), bottom=Side(style='thin'))
    
    current_month_str = datetime.utcnow().strftime("%Y-%m")
    
    row_num = 4
    for bot in bots:
        # Calculate Runs
        total_runs = len(bot.runs) if hasattr(bot, 'runs') and bot.runs else 0
        runs_this_month = sum(1 for r in bot.runs if r.started_on and r.started_on.startswith(current_month_str)) if hasattr(bot, 'runs') and bot.runs else 0
        hours_saved_this_month = runs_this_month * (bot.per_day_saving_hours or 0)
        
        dept_name = bot.department.name if bot.department else ""
        spoc_name = bot.spoc.name if bot.spoc else ""
        
        # Calculate availability
        desc_avail = "Yes" if bot.description and len(bot.description.strip()) > 0 else "No"
        pdd_avail = "Yes" if bot.pdd_link and len(bot.pdd_link.strip()) > 0 else "No"
        desc_text = bot.description if bot.description and len(bot.description.strip()) > 0 else "Description is not available"
        
        row_data = [
            row_num - 3, # S. No.
            dept_name or "-", # Category
            bot.use_case_name or "-", # CoBot Name
            bot.status or "-", # Status
            desc_avail, # Description Available
            pdd_avail, # PDD Available
            desc_text, # Description
            bot.hours_saved_monthly or 0, # Hours Saved (Monthly)
            hours_saved_this_month or 0, # Hours Saved (This Month)
            runs_this_month, # Runs This Month
            bot.comments or "-" # Remarks
        ]
        
        for col_num, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = val
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            
        row_num += 1
        
    # Footer (Signatures)
    row_num += 3
    
    # Left Signature
    ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=4)
    footer_left = ws.cell(row=row_num, column=2)
    footer_left.value = "Generated by"
    footer_left.font = Font(bold=True, color="7F7F7F")
    
    ws.merge_cells(start_row=row_num+1, start_column=2, end_row=row_num+1, end_column=4)
    footer_left_name = ws.cell(row=row_num+1, column=2)
    footer_left_name.value = "Adani CoBot System"
    footer_left_name.font = Font(size=14, bold=True, italic=True, color="1F4E78")
    
    # Right Signature
    ws.merge_cells(start_row=row_num, start_column=len(headers)-3, end_row=row_num, end_column=len(headers)-1)
    footer_right = ws.cell(row=row_num, column=len(headers)-3)
    footer_right.value = "Reviewed by"
    footer_right.font = Font(bold=True, color="7F7F7F")
    footer_right.alignment = Alignment(horizontal="right")
    
    ws.merge_cells(start_row=row_num+1, start_column=len(headers)-3, end_row=row_num+1, end_column=len(headers)-1)
    footer_right_name = ws.cell(row=row_num+1, column=len(headers)-3)
    footer_right_name.value = "Durgesh Tiwari"
    footer_right_name.font = Font(size=14, bold=True, italic=True, color="1F4E78")
    footer_right_name.alignment = Alignment(horizontal="right")
    
    # Save to stream
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    filename = f"CoBot_Report_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# Bot CRUD endpoints
@router.post("/bots")
def create_bot(bot_data: BotCreate, background_tasks: BackgroundTasks, db: Session = Depends(get_db), current_user: str = Depends(get_current_user_email)):
    """Create a new bot."""
    # Check if bot with same name already exists
    existing = db.query(Bot).filter(Bot.use_case_name == bot_data.use_case_name).first()
    if existing:
        raise HTTPException(status_code=400, detail="Bot with this use case name already exists")
        
    # Auto-generate use_case_no if not provided
    if not bot_data.use_case_no:
        last_bot = db.query(Bot).filter(Bot.use_case_no.like("AUC%")).order_by(Bot.id.desc()).first()
        if last_bot and last_bot.use_case_no:
            match = re.search(r'\d+', last_bot.use_case_no)
            if match:
                next_num = int(match.group()) + 1
                bot_data.use_case_no = f"AUC{next_num:03d}"
            else:
                bot_data.use_case_no = "AUC001"
        else:
            bot_data.use_case_no = "AUC001"
    
    new_bot = Bot(
        use_case_name=bot_data.use_case_name,
        use_case_no=bot_data.use_case_no,
        bot_name=bot_data.bot_name or bot_data.use_case_name,
        department_id=bot_data.department_id,
        spoc_id=bot_data.spoc_id,
        status=bot_data.status,
        developer=bot_data.developer,
        hours_saved_monthly=bot_data.hours_saved_monthly,
        pdd_link=bot_data.pdd_link,
        schedule_time=bot_data.schedule_time,
        description=bot_data.description,
        created_at=(datetime.utcnow() + timedelta(hours=5, minutes=30)).isoformat()
    )
    db.add(new_bot)
    db.commit()
    db.refresh(new_bot)
    
    # Fetch Department and SPOC names for the email
    dept_name = "N/A"
    if new_bot.department_id:
        dept = db.query(Department).filter(Department.id == new_bot.department_id).first()
        if dept:
            dept_name = dept.name
            
    spoc_name = "N/A"
    if new_bot.spoc_id:
        spoc = db.query(SPOC).filter(SPOC.id == new_bot.spoc_id).first()
        if spoc:
            spoc_name = spoc.name
            
    # Fetch admin emails
    admins = db.query(RegisteredUser).filter(RegisteredUser.role == "Admin").all()
    admin_emails = [admin.email for admin in admins if admin.email]
    
    if admin_emails:
        background_tasks.add_task(send_new_bot_notification, admin_emails, new_bot, dept_name, spoc_name)
        
    # Audit Log
    audit = AuditLog(
        entity_type="Bot",
        entity_id=str(new_bot.id),
        action_type="CREATE",
        new_value=json.dumps(bot_data.model_dump() if hasattr(bot_data, "model_dump") else bot_data.dict()),
        changed_by=current_user,
        timestamp=datetime.utcnow().isoformat()
    )
    db.add(audit)
    db.commit()
    
    return {"id": new_bot.id, "message": "Bot created successfully"}

@router.put("/bots/{bot_id}")
def update_bot(bot_id: int, bot_data: BotUpdate, db: Session = Depends(get_db), current_user: str = Depends(get_current_user_email)):
    """Update an existing bot."""
    bot = db.query(Bot).filter(Bot.id == bot_id).first()
    if not bot:
        raise HTTPException(status_code=404, detail="Bot not found")
        
    old_value = {
        "use_case_name": bot.use_case_name,
        "use_case_no": bot.use_case_no,
        "bot_name": bot.bot_name,
        "status": bot.status,
        "developer": bot.developer,
        "hours_saved_monthly": bot.hours_saved_monthly,
        "description": bot.description
    }
    
    # Update fields
    bot.use_case_name = bot_data.use_case_name
    bot.use_case_no = bot_data.use_case_no
    bot.bot_name = bot_data.bot_name or bot_data.use_case_name
    bot.department_id = bot_data.department_id if bot_data.department_id else bot.department_id
    bot.spoc_id = bot_data.spoc_id if bot_data.spoc_id else None
    bot.status = bot_data.status
    bot.developer = bot_data.developer
    bot.hours_saved_monthly = bot_data.hours_saved_monthly
    bot.pdd_link = bot_data.pdd_link
    bot.schedule_time = bot_data.schedule_time
    bot.description = bot_data.description
    
    # Audit Log
    audit = AuditLog(
        entity_type="Bot",
        entity_id=str(bot.id),
        action_type="UPDATE",
        old_value=json.dumps(old_value),
        new_value=json.dumps(bot_data.model_dump() if hasattr(bot_data, "model_dump") else bot_data.dict()),
        changed_by=current_user,
        timestamp=datetime.utcnow().isoformat()
    )
    db.add(audit)
    
    db.commit()
    return {"message": "Bot updated successfully"}

@router.delete("/bots/{bot_id}")
def delete_bot(bot_id: int, db: Session = Depends(get_db), current_user: str = Depends(get_current_user_email)):
    """Delete a bot."""
    bot = db.query(Bot).filter(Bot.id == bot_id).first()
    if not bot:
        raise HTTPException(status_code=404, detail="Bot not found")
        
    # Audit Log
    audit = AuditLog(
        entity_type="Bot",
        entity_id=str(bot.id),
        action_type="DELETE",
        old_value=json.dumps({"use_case_name": bot.use_case_name, "bot_name": bot.bot_name}),
        changed_by=current_user,
        timestamp=datetime.utcnow().isoformat()
    )
    db.add(audit)
    
    db.delete(bot)
    db.commit()
    return {"message": "Bot deleted successfully"}


# User Management endpoints
@router.get("/users", response_model=List[UserOut])
def get_users(db: Session = Depends(get_db)):
    """Fetch all registered users."""
    users = db.query(RegisteredUser).all()
    return users

@router.post("/users")
def add_user(user_data: UserCreate, db: Session = Depends(get_db), current_user: str = Depends(get_current_user_email)):
    """Register a new user."""
    existing = db.query(RegisteredUser).filter(RegisteredUser.email == user_data.email).first()
    if existing:
        raise HTTPException(status_code=400, detail="User already registered")
    
    new_user = RegisteredUser(
        email=user_data.email,
        name=user_data.name,
        role=user_data.role
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    
    # Audit Log
    audit = AuditLog(
        entity_type="User",
        entity_id=str(new_user.id),
        action_type="CREATE",
        new_value=json.dumps(user_data.model_dump() if hasattr(user_data, "model_dump") else user_data.dict()),
        changed_by=current_user,
        timestamp=datetime.utcnow().isoformat()
    )
    db.add(audit)
    db.commit()
    return {"message": "User registered successfully"}

@router.put("/users/{user_id}")
def update_user(user_id: int, user_data: UserUpdate, db: Session = Depends(get_db), current_user: str = Depends(get_current_user_email)):
    """Update user details or status."""
    user = db.query(RegisteredUser).filter(RegisteredUser.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
        
    old_value = {"email": user.email, "role": user.role, "is_active": user.is_active}
    
    user.email = user_data.email
    user.name = user_data.name
    user.role = user_data.role
    user.is_active = user_data.is_active
    
    # Audit Log
    audit = AuditLog(
        entity_type="User",
        entity_id=str(user.id),
        action_type="UPDATE",
        old_value=json.dumps(old_value),
        new_value=json.dumps(user_data.model_dump() if hasattr(user_data, "model_dump") else user_data.dict()),
        changed_by=current_user,
        timestamp=datetime.utcnow().isoformat()
    )
    db.add(audit)
    
    db.commit()
    return {"message": "User updated successfully"}

@router.delete("/users/{user_id}")
def delete_user(user_id: int, db: Session = Depends(get_db), current_user: str = Depends(get_current_user_email)):
    """Remove a registered user."""
    user = db.query(RegisteredUser).filter(RegisteredUser.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
        
    # Audit Log
    audit = AuditLog(
        entity_type="User",
        entity_id=str(user.id),
        action_type="DELETE",
        old_value=json.dumps({"email": user.email, "role": user.role}),
        changed_by=current_user,
        timestamp=datetime.utcnow().isoformat()
    )
    db.add(audit)
    
    db.delete(user)
    db.commit()
    return {"message": "User deleted successfully"}

@router.get("/check-user/{email}")
def check_user_registration(email: str, db: Session = Depends(get_db)):
    """Check if an email is registered and active."""
    user = db.query(RegisteredUser).filter(RegisteredUser.email == email, RegisteredUser.is_active == 1).first()
    if not user:
        return {"registered": False}
    return {"registered": True, "role": user.role}

@router.get("/audit-logs")
def get_audit_logs(db: Session = Depends(get_db)):
    """Fetch all admin activity audit logs."""
    logs = db.query(AuditLog).order_by(AuditLog.timestamp.desc()).all()
    return logs

