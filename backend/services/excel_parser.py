import openpyxl
from sqlalchemy.orm import Session
from typing import Tuple, List, Any
import re
from models import Department, SPOC, Bot, BotRun
from datetime import datetime, timedelta

def normalize_column_name(col: Any) -> str:
    """Normalize column names for consistent mapping."""
    if col is None:
        return ""
    col = str(col).lower().strip()
    col = re.sub(r'\s+', '_', col)
    col = re.sub(r'[^\w_]', '', col)
    return col

def get_sheet_data(sheet) -> List[dict]:
    """Helper to convert sheet to list of dicts based on header row."""
    rows = list(sheet.rows)
    if not rows:
        return []
    
    headers = [cell.value for cell in rows[0]]
    data = []
    
    for row in rows[1:]:
        row_data = {}
        has_data = False
        for i, cell in enumerate(row):
            if i < len(headers):
                val = cell.value
                # Normalize empty strings/spaces to None to match pandas behavior roughly
                if val is not None and isinstance(val, str) and not val.strip():
                    val = None
                
                if val is not None:
                    has_data = True
                
                row_data[headers[i]] = val
        
        if has_data:
            data.append(row_data)
            
    return data, headers

def parse_master_excel(file_path: str, db: Session) -> Tuple[int, int, int, int, List[str]]:
    """
    Parse the master bot list Excel file and store in database.
    Returns: (records_processed, departments_created, spocs_created, bots_created, errors)
    """
    errors = []
    records_processed = 0
    departments_created = 0
    spocs_created = 0
    bots_created = 0
    
    try:
        # Read Excel file
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        # Process ONLY Sheet1 (or single main sheet)
        if 'Sheet1' in wb.sheetnames:
            target_sheets = ['Sheet1']
        else:
            # Fallback to first sheet
            target_sheets = [wb.sheetnames[0]]
            
        # Iterate through target sheets only
        for sheet_name in target_sheets:
            sheet = wb[sheet_name]
            
            # Check if this sheet looks like a bot list
            rows = list(sheet.rows)
            if not rows:
                continue
                
            header_row = rows[0]
            col_map = {} # normalized -> current_header_name
            header_index_map = {} # normalized -> column_index (0-based)
            
            # Create a quick map to check validation
            for idx, cell in enumerate(header_row):
                val = cell.value
                if val:
                    normalized = normalize_column_name(val)
                    col_map[normalized] = val
                    header_index_map[normalized] = idx
            
            # Check for critical column "Use Case Name" or similar
            has_use_case = False
            for k in ['use_case_name_/_bot_name', 'use_case_name', 'bot_name', 'usecase_name']:
                if k in header_index_map:
                    has_use_case = True
                    break
            
            if not has_use_case:
                # Skip sheets that aren't bot lists (e.g. "Summary", "Dropdowns")
                continue
            
            # Process this sheet
            for row_idx, row in enumerate(rows[1:], start=2):
                try:
                    # Helper to safely get value by normalized column name
                    def get_val(norm_key):
                        if norm_key in header_index_map:
                            col_idx = header_index_map[norm_key]
                            if col_idx < len(row):
                                val = row[col_idx].value
                                if val is not None and isinstance(val, str) and not val.strip():
                                    val = None
                                return val
                        return None

                    # Get use case name
                    use_case_name = None
                    for key in ['use_case_name_/_bot_name', 'use_case_name', 'bot_name', 'usecase_name']:
                        val = get_val(key)
                        if val is not None:
                            use_case_name = str(val).strip()
                            break
                    
                    if not use_case_name or use_case_name.lower() == 'nan':
                        continue
                    
                    records_processed += 1
                    
                    # Get or create department
                    dept_name = None
                    val = get_val('department')
                    if val is not None:
                        dept_name = str(val).strip()
                    
                    department = None
                    if dept_name and dept_name.lower() != 'nan':
                        department = db.query(Department).filter(Department.name == dept_name).first()
                        if not department:
                            department = Department(name=dept_name)
                            db.add(department)
                            db.flush()
                            departments_created += 1
                    
                    # Get or create SPOC
                    spoc_name = None
                    val = get_val('business_spoc')
                    if val is not None:
                        spoc_name = str(val).strip()
                    
                    spoc = None
                    if spoc_name and spoc_name.lower() != 'nan':
                        spoc = db.query(SPOC).filter(SPOC.name == spoc_name).first()
                        if not spoc:
                            spoc = SPOC(name=spoc_name)
                            db.add(spoc)
                            db.flush()
                            spocs_created += 1
                    
                    # Check if bot already exists
                    existing_bot = db.query(Bot).filter(Bot.use_case_name == use_case_name).first()
                    
                    # Helper to get string val
                    def get_str(key):
                        v = get_val(key)
                        if v is not None:
                            return str(v).strip()
                        return None

                    # Get other fields
                    status = None
                    for key in ['status', 'bot_status', 'state', 'current_status', 'status_of_bot', 'status_of_automation']:
                        val = get_str(key)
                        if val:
                             status = val
                             break
                    if not status:
                        status = 'Unknown'
                    developer = get_str('developer')
                    
                    hours_saved = get_val('hours_optimization_per_month')
                    if hours_saved is None:
                        hours_saved = 0
                    else:
                        try:
                            hours_saved = float(hours_saved)
                        except:
                            hours_saved = 0
                    
                    deployed_date = get_str('deployed_date')
                    if deployed_date and isinstance(get_val('deployed_date'), datetime):
                         deployed_date = get_val('deployed_date').strftime('%Y-%m-%d')
                         
                    schedule = get_str('schedule')
                    # Prioritize 'Description' column, fallback to 'Comments'
                    description = None
                    for key in ['description', 'bot_description', 'use_case_description', 'comments_/_pending_actions']:
                        val = get_str(key)
                        if val:
                             description = val
                             break
                    bu_name = get_str('bu_name')
                    team = get_str('team')
                    pdd_location = get_str('pdd_location_/_release_notes_location')
                    
                    # Get PDD Link from Link column (clickable URL)
                    pdd_link = get_str('link')
                    
                    # Get Use Case No
                    use_case_no = get_str('use_case_no')
                    
                    # Get Schedule Time IN Runner Machine
                    schedule_time = get_str('schedule_time_in_runner_machine')
                    
                    # New fields
                    sr_no_val = get_val('srno')
                    sr_no = int(sr_no_val) if sr_no_val and str(sr_no_val).isdigit() else None
                    
                    start_date = get_str('startdate')
                    if start_date and isinstance(get_val('startdate'), datetime):
                         start_date = get_val('startdate').strftime('%Y-%m-%d')
                         
                    end_date = get_str('enddate')
                    if end_date and isinstance(get_val('enddate'), datetime):
                         end_date = get_val('enddate').strftime('%Y-%m-%d')

                    comments = get_str('comments_/_pending_actions')
                    machine_ip = get_str('botrunnermachine/_ip')
                    execution_time = get_str('execution_time')
                    bot_running_status = get_str('bot_run_status')
                    last_run_status = get_str('bot_run_status')
                    key_benefits = get_str('key_benefits')
                    bot_name = get_str('use_case_name_/_bot_name') or use_case_name
                    
                    # New calculation logic fields
                    frequency = get_str('frequency')  # Col 24
                    schedule_details = get_str('schedule_details')  # Col 26
                    
                    per_day_saving = get_val('per_day_saving_hours')  # Col 17
                    if per_day_saving is None:
                        per_day_saving = 0
                    else:
                        try:
                            # Handle string formats like "10 min", "2 hr"
                            val_str = str(per_day_saving).lower().strip()
                            if 'min' in val_str:
                                # Extract number and divide by 60
                                import re
                                num = float(re.findall(r"[\d\.]+", val_str)[0])
                                per_day_saving = num / 60.0
                            elif 'hr' in val_str:
                                # Extract number
                                import re
                                num = float(re.findall(r"[\d\.]+", val_str)[0])
                                per_day_saving = num
                            else:
                                per_day_saving = float(per_day_saving)
                        except:
                            per_day_saving = 0


                    if existing_bot:
                        # Update existing bot
                        existing_bot.bot_name = bot_name
                        existing_bot.status = status
                        existing_bot.developer = developer
                        existing_bot.hours_saved_monthly = hours_saved
                        existing_bot.deployed_date = deployed_date
                        existing_bot.schedule = schedule
                        existing_bot.description = description
                        existing_bot.bu_name = bu_name
                        existing_bot.team = team
                        existing_bot.pdd_location = pdd_location
                        
                        # New fields update
                        existing_bot.sr_no = sr_no
                        existing_bot.start_date = start_date
                        existing_bot.end_date = end_date
                        existing_bot.comments = comments
                        existing_bot.machine_ip = machine_ip
                        existing_bot.execution_time = execution_time
                        existing_bot.bot_running_status = bot_running_status
                        existing_bot.last_run_status = last_run_status
                        existing_bot.key_benefits = key_benefits
                        existing_bot.pdd_link = pdd_link
                        existing_bot.use_case_no = use_case_no
                        existing_bot.schedule_time = schedule_time
                        
                        # Calculation logic fields
                        existing_bot.frequency = frequency
                        existing_bot.schedule_details = schedule_details
                        existing_bot.per_day_saving_hours = per_day_saving

                        if department:
                            existing_bot.department_id = department.id
                        if spoc:
                            existing_bot.spoc_id = spoc.id
                    else:
                        # Create new bot
                        bot = Bot(
                            use_case_name=use_case_name,
                            bot_name=bot_name,
                            department_id=department.id if department else None,
                            spoc_id=spoc.id if spoc else None,
                            status=status,
                            developer=developer,
                            hours_saved_monthly=hours_saved,
                            deployed_date=deployed_date,
                            schedule=schedule,
                            description=description,
                            bu_name=bu_name,
                            team=team,
                            pdd_location=pdd_location,
                            created_at=(datetime.utcnow() + timedelta(hours=5, minutes=30)).isoformat(),
                            
                            # New fields
                            sr_no=sr_no,
                            start_date=start_date,
                            end_date=end_date,
                            comments=comments,
                            machine_ip=machine_ip,
                            execution_time=execution_time,
                            bot_running_status=bot_running_status,
                            last_run_status=last_run_status,
                            key_benefits=key_benefits,
                            pdd_link=pdd_link,
                            use_case_no=use_case_no,
                            schedule_time=schedule_time,
                            
                            # Calculation logic fields
                            frequency=frequency,
                            schedule_details=schedule_details,
                            per_day_saving_hours=per_day_saving
                        )

                        db.add(bot)
                        bots_created += 1
                    
                except Exception as e:
                    errors.append(f"Sheet {sheet_name} Row {row_idx}: {str(e)}")
            
            db.commit()
        
    except Exception as e:
        errors.append(f"File processing error: {str(e)}")
        db.rollback()
    
    return records_processed, departments_created, spocs_created, bots_created, errors


def parse_daily_report(file_path: str, db: Session, report_date: str = None) -> Tuple[int, int, float, List[str]]:
    """
    Parse the daily bot status report and store run records.
    Returns: (runs_processed, unique_bots_matched, hours_saved, errors)
    """
    errors = []
    runs_processed = 0
    bots_matched = 0
    matched_bot_ids = set()
    
    if not report_date:
        # Use IST (UTC+5:30)
        ist_now = datetime.utcnow() + timedelta(hours=5, minutes=30)
        report_date = ist_now.strftime('%Y-%m-%d')
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        # Look for Control Room Dump sheet or Bot Status
        sheet_name = None
        for name in wb.sheetnames:
            lower_name = name.lower()
            if 'control room' in lower_name or 'dump' in lower_name or 'daily bot status' in lower_name:
                sheet_name = name
                break
        
        # Fallback: Look for just "Status" if nothing else
        if not sheet_name:
             for name in wb.sheetnames:
                 if 'status' in name.lower():
                     sheet_name = name
                     break
        
        if not sheet_name:
            sheet_name = wb.sheetnames[0]
            
        sheet = wb[sheet_name]
        
        # Create column mapping
        rows = list(sheet.rows)
        if not rows:
             return 0, 0, 0.0, ["Empty file"]

        header_row = rows[0]
        col_map = {} 
        header_index_map = {}
        
        for idx, cell in enumerate(header_row):
            val = cell.value
            if val:
                normalized = normalize_column_name(val)
                col_map[normalized] = val
                header_index_map[normalized] = idx
        
        # Validation: Check for required columns
        # Based on dump file analysis or assumption of critical fields
        required_partial_keys = ['activity_name', 'status', 'started_on', 'ended_on']
        missing_cols = []
        for req in required_partial_keys:
            found = False
            for key in header_index_map:
                if req in key:
                    found = True
                    break
            if not found:
                 # Check for alternate names
                 if req == 'activity_name':
                     if any(k in header_index_map for k in ['automation_name', 'bot_name']):
                         found = True
            
            if not found:
                missing_cols.append(req)
        
        if missing_cols:
            return 0, 0, 0.0, [f"Missing required columns: {', '.join(missing_cols)}"]
        
        # Delete existing runs for this report date
        db.query(BotRun).filter(BotRun.report_date == report_date).delete()
        
        for row_idx, row in enumerate(rows[1:], start=2):
            try:
                def get_val(norm_key):
                    if norm_key in header_index_map:
                        col_idx = header_index_map[norm_key]
                        if col_idx < len(row):
                            val = row[col_idx].value
                            if val is not None and isinstance(val, str) and not val.strip():
                                return None
                            return val
                    return None

                # Get activity name
                activity_name = None
                for key in ['activity_name', 'automation_name', 'bot_name']:
                    val = get_val(key)
                    if val is not None:
                        activity_name = str(val).strip()
                        break
                
                if not activity_name or activity_name.lower() == 'nan':
                    continue
                
                # Try to match with existing bot
                # 1. Exact match by use_case_name
                bot = db.query(Bot).filter(Bot.use_case_name == activity_name).first()
                
                if not bot:
                    # 2. Exact match by use_case_no (e.g. "AUC004")
                    bot = db.query(Bot).filter(Bot.use_case_no == activity_name).first()

                if not bot:
                    # 3. Fuzzy match: Check if activity_name starts with any bot's use_case_no
                    # This handles "AUC004 AGEL..." if the bot use_case_no is "AUC004"
                    bots_with_no = db.query(Bot).filter(Bot.use_case_no != None).all()
                    for b in bots_with_no:
                        # Ensure use_case_no is specific enough (at least 3 chars)
                        if b.use_case_no and len(b.use_case_no) >= 3 and activity_name.startswith(b.use_case_no):
                            bot = b
                            break
                
                if not bot:
                    # 4. STRICTER fuzzy match by base name (must be exact match before version)
                    # This handles "BotName.1" matching "BotName" 
                    # But avoids "BotNameExtra" matching "BotName"
                    base_name = re.split(r'\.\d+', activity_name)[0]
                    # Check if base_name is exactly equal to a bot name
                    bot = db.query(Bot).filter(Bot.use_case_name == base_name).first()

                # REMOVED RISKY MATCHING STEPS (5 & 6) 
                # Prefix matching and substring matching were causing false positives.

                
                if not bot:
                    continue
                
                runs_processed += 1
                if bot.id not in matched_bot_ids:
                    bots_matched += 1
                    matched_bot_ids.add(bot.id)
                
                # Get run status
                run_status = get_val('status')
                if run_status is not None:
                    run_status = str(run_status).strip()
                else:
                    run_status = 'Unknown'
                
                # Get timing info
                started_on = get_val('started_on')
                if started_on is not None:
                    started_on = str(started_on)
                else:
                    started_on = None
                
                ended_on = get_val('ended_on')
                if ended_on is not None:
                    ended_on = str(ended_on)
                else:
                    ended_on = None
                
                device_name = get_val('device_name')
                if device_name is not None:
                    device_name = str(device_name).strip()
                else:
                    device_name = None
                
                automation_type = get_val('automation_type')
                if automation_type is not None:
                    automation_type = str(automation_type).strip()
                else:
                    automation_type = None
                
                # Create bot run record
                bot_run = BotRun(
                    bot_id=bot.id,
                    run_status=run_status,
                    started_on=started_on,
                    ended_on=ended_on,
                    device_name=device_name,
                    report_date=report_date,
                    automation_type=automation_type
                )
                db.add(bot_run)
                
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")
        
        db.commit()
        
    except Exception as e:
        errors.append(f"File processing error: {str(e)}")
        db.rollback()
    
    
    # Calculate unique hours saved using new logic (Runs * Value)
    total_hours_saved = 0.0
    
    # Re-query runs
    runs = db.query(BotRun).filter(BotRun.report_date == report_date).all()
    bot_run_counts = {}
    
    for run in runs:
        if run.bot_id and run.run_status and 'completed' in str(run.run_status).lower():
            bot_run_counts[run.bot_id] = bot_run_counts.get(run.bot_id, 0) + 1
            
    for bot_id, count in bot_run_counts.items():
         bot = db.query(Bot).get(bot_id)
         if bot:
             # Inline calculation logic to avoid circular import
             monthly = bot.hours_saved_monthly or 0
             per_day = bot.per_day_saving_hours or 0
             sched = (bot.schedule or "").lower()
             freq = float(bot.frequency) if bot.frequency else 1.0
             if freq <= 0: freq = 1.0
             
             if 'on demand' in sched or 'multiple' in sched:
                 total_hours_saved += (per_day * count)
             else:
                 # Daily/Weekly/Monthly
                 val_per_run = monthly / freq
                 total_hours_saved += (val_per_run * count)
             
    return runs_processed, len(matched_bot_ids), total_hours_saved, errors
