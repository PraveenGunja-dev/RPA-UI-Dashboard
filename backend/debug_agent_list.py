import openpyxl
import os
import sys

path = r"d:\Adani_projects\RPA_Dashboard\data\Agent-list.xlsx"

if not os.path.exists(path):
    print(f"File not found: {path}")
    sys.exit(1)

print(f"File size: {os.path.getsize(path)}")

try:
    wb = openpyxl.load_workbook(path, data_only=True)
    print("Success loading workbook")
    print("Sheets:", wb.sheetnames)
    
    # Print first row of headers from active sheet
    sheet = wb.active
    rows = list(sheet.iter_rows(max_row=1, values_only=True))
    if rows:
        print("Headers:", rows[0])
    
except Exception as e:
    print(f"Failed: {e}")
